# All the imports
import requests
from bs4 import BeautifulSoup
import os
import re
import pandas as pd
from nltk.tokenize import TweetTokenizer

# Extracting information from the 'Input.xlsx' in the format [{'URL_ID':String,'URL:String}]
from openpyxl import load_workbook

workbook = load_workbook(filename='Input.xlsx')

sheet = workbook.active

URLs=[]

for row in sheet.iter_rows(min_row=2,values_only=True):
    URLs.append({"URL_ID":row[0],"URL":row[1]})

# Counts number of syllables
def count_syllables(word):
    word = word.lower()
    vowels = "aeiouy"
    syllable_count = 0

    # Each group of vowels counts as one syllable
    syllable_count += len([char for i, char in enumerate(word) if char in vowels and (i == 0 or word[i-1] not in vowels)])

    # Words ending in 'e' usually don't increase the syllable count, unless the 'e' is the only vowel
    if word.endswith('e') and syllable_count > 1:
        syllable_count -= 1

    # Words ending in 'le' or 'les' usually increase the syllable count
    if word.endswith(('le', 'les')) and len(word) > 2 and word[-3] not in vowels:
        syllable_count += 1

    # Words ending in 'ed' but not 'eed' usually don't increase the syllable count
    if word.endswith('ed') and not word.endswith('eed'):
        syllable_count -= 1

    # If no syllables have been counted, count one
    if syllable_count == 0:
        syllable_count += 1

    return syllable_count

# Returns a list that contains stopwords
def Stopwords():
  path_stopwords='C:/Users/dell/Desktop/Sentiment Analysis/StopWords'
  stopwords=[]
  files=os.listdir(path_stopwords)
  for file in files:
    with open(os.path.join(path_stopwords,file),'r',encoding="latin1") as f:
       sw=f.read().splitlines()
       for s in sw:
         x=s.split('|')[0]
         stopwords.append(x)
  stopwords=[word.lower() for word in stopwords]
  return stopwords

# Preprocessing raw html to extract info like input words, sentence count,etc
def Preprocessing(url):
    page=requests.get(url['URL'])
    soup=BeautifulSoup(page.text,'html')

    content = soup.find('div', class_='td-post-content')

    if content:

      # Extract text from <p> and <li> tags
      html_contents = content.find_all(['p','li'])

      paragraphs = [html_content.text for html_content in html_contents]

      single_string_paragraph=["".join(paragraphs)]
      sentence_split=single_string_paragraph[0].split(".")
      sentence_count=len(sentence_split)-1

      tokenizer = TweetTokenizer()
      sentence_tokenized =[tokenizer.tokenize(sentence) for sentence in sentence_split]
      sentence_tokenized=sentence_tokenized[:-1]
      sentence_tokenized = [[word.lower() for word in sentence if word.isalnum() or re.match(r"[\w']",word)] for sentence in sentence_tokenized ]
      word_count=sum([len(sentence) for sentence in sentence_tokenized])

      sw=Stopwords()
      input_words=[word for sentence in sentence_tokenized for word in sentence if word not in sw ]

      return (input_words,sentence_count,word_count,sentence_tokenized)
    else:
      return None

# Returns a tuple that contains positive and negative words
def PositiveNegativeWords(input_words):
    pw_file_path="positive-words.txt"
    nw_file_path="negative-words.txt"

    with open(pw_file_path,'r') as file:
      content=file.read()
      positive_words=content.splitlines()

    with open(nw_file_path,'r',encoding="latin1") as file:
      content=file.read()
      negative_words=content.splitlines()

    return (positive_words,negative_words)

# Returns the metrics like positive score, polarity,etc
def Metrics(input_words,positive_words,negative_words,sentence_count,sentence_tokenized,word_count,url):
  # Positive Score
  intersecting_pw_words=[word for word in input_words if word in positive_words]
  positive_score=len(intersecting_pw_words)

  # Negative Score
  intersecting_nw_words=[word for word in input_words if word in negative_words]
  negative_score=len(intersecting_nw_words)

  # Polarity Score
  polarity_score=(positive_score-negative_score)/(positive_score+negative_score)+0.000001

  # Subjective Score
  subjective_score=(positive_score+negative_score)/(sentence_count)+0.000001

  # Average Sentence Length
  avg_sen_length=word_count/sentence_count

  # Percentage of Complex Words
  syllable_count=[count_syllables(word) for sentence in sentence_tokenized for word in sentence]
  # Complex Word Count
  complex_word_count=0
  for count in syllable_count:
    if(count>2):
      complex_word_count+=1
      per_comp_words=complex_word_count/word_count

  # Fog Index
  fog_index=0.4*(avg_sen_length+per_comp_words)

  # Average Number of Words Per Sentence
  avg_no_words_sen=avg_sen_length

  # Word Count
  wc=len(input_words)

  # Syllable Per Word
  syll_per_word=sum(syllable_count)/word_count

  # Average Word Length
  total_alphabet=[len(word) for sentence in sentence_tokenized for word in sentence]
  avg_word_length=sum(total_alphabet)/word_count

  # Personal Pronouns
  personal_pronouns=['i','we','my','ours','us']
  strings_to_remove=['Us','US']
  sentence_array=[ word for sentence in sentence_tokenized for word in sentence]
  sentence_array=[word for word in sentence_array if word not in strings_to_remove]
  personal_pronouns_present=[word for word in sentence_array if word in personal_pronouns]
  per_pronouns=len(personal_pronouns_present)

  return {'URL_ID':url['URL_ID'],'URL':url['URL'],'POSITIVE SCORE':positive_score,'NEGATIVE SCORE':negative_score,
                'POLARITY SCORE':polarity_score,'SUBJECTIVE SCORE':subjective_score,'AVG SENTENCE LENGTH':avg_sen_length,
                'PERCENTAGE OF COMPLEX WORDS':per_comp_words,'FOG INDEX':fog_index,'AVG NUMBER OF WORDS PER SENTENCE':avg_no_words_sen,
                'COMPLEX WORD COUNT':complex_word_count,'WORD COUNT':word_count,'SYLLABLE PER WORD':syll_per_word,'PERSONAL PRONOUNS':
                per_pronouns,'AVG WORD LENGTH':avg_word_length
                }
       

# Generates the excel file that is the output
def Generate_excel(urls):
    site_metrics=[]
    for url in urls:
        if Preprocessing(url)==None:
          site_metrics.append(
              {'URL_ID':url['URL_ID'],'URL':url['URL'],'POSITIVE SCORE':'NULL','NEGATIVE SCORE':'NULL',
                'POLARITY SCORE':'NULL','SUBJECTIVE SCORE':'NULL','AVG SENTENCE LENGTH':'NULL',
                'PERCENTAGE OF COMPLEX WORDS':'NULL','FOG INDEX':'NULL','AVG NUMBER OF WORDS PER SENTENCE':'NULL',
                'COMPLEX WORD COUNT':'NULL','WORD COUNT':'NULL','SYLLABLE PER WORD':'NULL','PERSONAL PRONOUNS':
                'NULL','AVG WORD LENGTH':'NULL'
                }
          )
        else:
          input_words,sentence_count,word_count,sentence_tokenized=Preprocessing(url)
          positive_words,negative_words=PositiveNegativeWords(input_words)
          site_metrics.append(Metrics(input_words,positive_words,negative_words,sentence_count,sentence_tokenized,word_count,url))

    df = pd.DataFrame(site_metrics)
    excel_file = "Output.xlsx"
    df.to_excel(excel_file, index=False)
    print(f"Excel sheet '{excel_file}' has been generated.")

Generate_excel(URLs)